"""Streaming Excel reader.

Two passes per sheet, both bounded by the *true* data extent rather than the
declared dimension (stray formatting routinely inflates the declared range to
XFD1048576):

1. A streaming XML parse for merges, hidden rows/columns, formula cells and
   the real extent (:func:`sheet_structure`).
2. openpyxl read-only value streaming into a trimmed grid (:func:`read_grid`).

Ported from ``reference/excel_loader.py`` (the frozen prototype); the value
side lands in plain Python lists here — polars typing happens in
``extract.py``.
"""

from __future__ import annotations

import logging
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from xl2y.errors import SheetNotFoundError, UnsupportedFormatError

logger = logging.getLogger(__name__)

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _is_empty(v: Any) -> bool:
    return v is None or str(v).strip() == ""


def check_format(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix in SUPPORTED_SUFFIXES:
        return
    if suffix in {".xls", ".xlsb"}:
        raise UnsupportedFormatError(
            f"{path.name}: {suffix} files cannot be read by openpyxl. "
            "Convert to .xlsx first (Excel: Save As; or `soffice --convert-to "
            "xlsx`; or pandas.read_excel with engine='xlrd'/'pyxlsb')."
        )
    raise UnsupportedFormatError(
        f"{path.name}: unrecognised extension {suffix!r}; expected one of "
        f"{sorted(SUPPORTED_SUFFIXES)}."
    )


def _localname(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def sheet_meta(path: Path) -> dict[str, tuple[str, bool]]:
    """Sheet order/type without loading the workbook.

    Returns an ordered mapping ``{sheet_name: (zip_member_path, is_chartsheet)}``
    parsed from ``xl/workbook.xml`` and its relationships.
    """
    RELS_NS = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    with zipfile.ZipFile(path) as zf:
        rels: dict[str, tuple[str, str]] = {}  # rId -> (target, type)
        root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        for rel in root:
            rels[rel.get("Id")] = (rel.get("Target"), rel.get("Type", ""))
        meta: dict[str, tuple[str, bool]] = {}
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        for el in wb_root.iter():
            if _localname(el.tag) != "sheet":
                continue
            rid = el.get(f"{{{RELS_NS}}}id")
            target, rel_type = rels.get(rid, ("", ""))
            # Targets may be absolute ("/xl/worksheets/sheet1.xml") or
            # relative to xl/ ("worksheets/sheet1.xml").
            member = (
                target.lstrip("/")
                if target.startswith("/")
                else f"xl/{target}"
            )
            is_chart = "chartsheet" in rel_type or "chartsheets/" in member
            meta[el.get("name")] = (member, is_chart)
        return meta


def _require_sheet(meta: dict[str, tuple[str, bool]], name: str) -> None:
    if name not in meta:
        raise SheetNotFoundError(
            f"Sheet {name!r} not found. Available: {list(meta)}"
        )
    if meta[name][1]:
        raise ValueError(
            f"Sheet {name!r} is a chartsheet and contains no cell data."
        )


@dataclass
class RawSheet:
    """Everything the extractor needs about one worksheet, fully in memory
    but bounded by the true data extent (never the declared dimension)."""

    name: str
    grid: list[list]  # trimmed rows; [] = empty row
    merged: list[tuple[int, int, int, int]]  # (min_r, min_c, max_r, max_c) 1-idx
    hidden_rows: set[int]  # 0-indexed
    hidden_col_intervals: list[tuple[int, int]]  # 0-indexed inclusive
    formulas: list[tuple[int, int]]  # (row, col) 1-indexed


def sheet_structure(
    zf: zipfile.ZipFile, member: str
) -> tuple[
    list[tuple[int, int, int, int]],
    set[int],
    list[tuple[int, int]],
    list[tuple[int, int]],
    tuple[int, int],
]:
    """Merged ranges + hidden rows/columns via a streaming XML parse.

    Reading these through a full ``read_only=False`` workbook load would
    materialise every cell object -- catastrophic on dimension-bloated
    files -- when all we need is three kinds of small elements.

    Returns
    -------
    merged_ranges : list of (min_row, min_col, max_row, max_col), 1-indexed
    hidden_rows   : 0-indexed row numbers
    hidden_cols   : list of 0-indexed (min, max) inclusive intervals
    formula_cells : list of (row, col), 1-indexed
    data_extent   : (max_row, max_col), 1-indexed; (0, 0) if no cell holds
                    a value.
    """
    merged: list[tuple[int, int, int, int]] = []
    hidden_rows: set[int] = set()
    hidden_col_intervals: list[tuple[int, int]] = []
    formulas: list[tuple[int, int]] = []
    max_row = max_col = 0
    truthy = {"1", "true"}
    # Cell/row `r` attributes are optional in OOXML; some writers omit them
    # and imply position, so track counters as we stream.
    cur_row = 0
    cur_col = 0
    cur_cell: tuple[int, int] | None = None
    with zf.open(member) as fh:
        for event, el in ET.iterparse(fh, events=("start", "end")):
            tag = _localname(el.tag)
            if event == "start":
                if tag == "row":
                    cur_row = int(el.get("r")) if el.get("r") else cur_row + 1
                    cur_col = 0
                elif tag == "c":
                    ref = el.get("r")
                    if ref:
                        cur_row, cur_col = coordinate_to_tuple(ref)
                    else:
                        cur_col += 1
                    cur_cell = (cur_row, cur_col)
                continue
            # end events
            if tag == "f":
                if cur_cell is not None:
                    formulas.append(cur_cell)
            elif tag in ("v", "is"):  # cell actually holds a value
                if cur_row > max_row:
                    max_row = cur_row
                if cur_col > max_col:
                    max_col = cur_col
            elif tag == "mergeCell":
                min_c, min_r, max_c, max_r = range_boundaries(el.get("ref"))
                merged.append((min_r, min_c, max_r, max_c))
            elif tag == "row":
                if el.get("hidden") in truthy:
                    hidden_rows.add(cur_row - 1)
                el.clear()  # rows dominate the file; free as we stream
            elif tag == "col":
                if el.get("hidden") in truthy:
                    hidden_col_intervals.append(
                        (int(el.get("min")) - 1, int(el.get("max")) - 1)
                    )
    return (
        merged,
        hidden_rows,
        hidden_col_intervals,
        formulas,
        (max_row, max_col),
    )


def read_grid(ws: Any, extent: tuple[int, int]) -> list[list[Any]]:
    """Stream cell values (read_only worksheet) into a trimmed grid.

    ``reset_dimensions()`` has been called, so iteration ends at the true
    data extent regardless of the declared dimension. Interior empty rows
    are yielded as placeholders (Excel row numbers are preserved); each row
    is trimmed of trailing empty cells, and runs of fully-empty rows are
    only flushed when a later non-empty row appears.
    """
    max_row, max_col = extent
    if max_row == 0:
        return []
    grid: list[list[Any]] = []
    pending_empty = 0
    for row in ws.iter_rows(max_row=max_row, max_col=max_col, values_only=True):
        vals = list(row)
        while vals and _is_empty(vals[-1]):
            vals.pop()
        if not vals:
            pending_empty += 1
        else:
            if pending_empty:
                grid.extend([] for _ in range(pending_empty))
                pending_empty = 0
            grid.append(vals)
    return grid  # trailing empty rows deliberately dropped


def _build_raw(ws: Any, zf: zipfile.ZipFile, member: str, name: str) -> RawSheet:
    # Don't trust the declared dimension (stray formatting can inflate it to
    # XFD1048576); let openpyxl derive the extent from the actual data.
    ws.reset_dimensions()
    merged, hidden_rows, hidden_cols, formulas, extent = sheet_structure(
        zf, member
    )
    grid = read_grid(ws, extent)
    return RawSheet(
        name=name,
        grid=grid,
        merged=merged,
        hidden_rows=hidden_rows,
        hidden_col_intervals=hidden_cols,
        formulas=formulas,
    )


def read_sheet(path: str | Path, sheet_name: str) -> RawSheet:
    """Read one worksheet into a :class:`RawSheet`."""
    path = Path(path)
    check_format(path)
    meta = sheet_meta(path)
    _require_sheet(meta, sheet_name)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        with zipfile.ZipFile(path) as zf:
            return _build_raw(
                wb[sheet_name], zf, meta[sheet_name][0], sheet_name
            )
    finally:
        wb.close()


def read_all(path: str | Path) -> Iterator[RawSheet]:
    """Yield one :class:`RawSheet` per worksheet, one at a time.

    Chartsheets are skipped (with an INFO log). Yielding lazily keeps peak
    memory at the largest single sheet, not the whole workbook.
    """
    path = Path(path)
    check_format(path)
    meta = sheet_meta(path)
    skipped = [n for n, (_, chart) in meta.items() if chart]
    if skipped:
        logger.info("Skipping non-worksheet tab(s): %s", skipped)
    names = [n for n, (_, chart) in meta.items() if not chart]
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        with zipfile.ZipFile(path) as zf:
            for name in names:
                yield _build_raw(wb[name], zf, meta[name][0], name)
    finally:
        wb.close()
