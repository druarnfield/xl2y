from openpyxl import load_workbook

from tests import fixtures


def test_simple_book(tmp_path):
    p = fixtures.simple_book(tmp_path / "s.xlsx")
    ws = load_workbook(p)["Data"]
    assert ws["A1"].value == "Store"
    assert ws.max_row == 4


def test_nasty_book_has_banner_merge(tmp_path):
    p = fixtures.nasty_book(tmp_path / "n.xlsx")
    ws = load_workbook(p)["Report"]
    assert any(str(r) == "A1:D1" for r in ws.merged_cells.ranges)
