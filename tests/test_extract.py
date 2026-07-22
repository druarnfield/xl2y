import polars as pl
import pytest

from tests import fixtures
from xl2y import extract, reader
from xl2y.errors import EmptySheetError


def _extracted(tmp_path, builder, sheet, **opts):
    p = builder(tmp_path / "f.xlsx")
    return extract.extract_table(reader.read_sheet(p, sheet), **opts)


def test_simple_table(tmp_path):
    ex = _extracted(tmp_path, fixtures.simple_book, "Data")
    assert ex.df.columns == ["Store", "Date", "Revenue"]
    assert ex.df.height == 3
    assert ex.excel_rows == [2, 3, 4]


def test_banner_becomes_comment_not_header(tmp_path):
    ex = _extracted(tmp_path, fixtures.nasty_book, "Report")
    assert ex.df.columns == ["Store", "Date", "Units", "Revenue"]
    assert ex.comments[0]["kind"] == "merged"
    assert "CONFIDENTIAL" in ex.comments[0]["text"]


def test_sparse_rows_default_comment(tmp_path):
    ex = _extracted(tmp_path, fixtures.nasty_book, "Report")
    texts = [c["text"] for c in ex.comments if c["kind"] == "sparse"]
    assert "Northern Region" in texts
    assert ex.df.height == 3  # section rows stripped


def test_sparse_rows_section_mode(tmp_path):
    ex = _extracted(
        tmp_path, fixtures.nasty_book, "Report", sparse_rows="section"
    )
    assert ex.df.columns[0] == "section"
    assert ex.df["section"].to_list() == [
        "Northern Region",
        "Northern Region",
        "Southern Region",
    ]


def test_multi_row_header(tmp_path):
    ex = _extracted(
        tmp_path, fixtures.multi_header_book, "Wide", header_rows=2
    )
    assert ex.df.columns == ["Store", "Revenue Q1", "Revenue Q2"]


def test_merged_banner_kept_inside_multi_row_header(tmp_path):
    # A wide merged banner in the top row is stripped as a title under the
    # default single-row header, but when the caller asks for a multi-row
    # header it belongs to that header — it must not be pulled out as a
    # comment and shift the header down onto a data row.
    from openpyxl import Workbook

    p = tmp_path / "banner.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"
    ws["A1"] = "FY25/26 Tracker"
    ws.merge_cells("A1:D1")  # 1: full-width banner
    ws.append(["Region", "Q1", "Q2", "Q3"])  # 2: header
    ws.append([None, "gbp", "gbp", "gbp"])  # 3: units sub-header
    ws.append(["Alpha", 1, 2, 3])  # 4: data
    ws.append(["Beta", 4, 5, 6])  # 5: data
    wb.save(p)

    ex = extract.extract_table(reader.read_sheet(p, "All Data"), header_rows=3)
    # The banner stays in the header, so data begins at Excel row 4 (not 5)
    # and the banner is not recorded as a comment.
    assert ex.excel_rows == [4, 5]
    assert ex.df.height == 2
    assert not [c for c in ex.comments if c["kind"] == "merged"]
    assert ex.df.columns[0] == "FY25/26 Tracker Region"


def test_header_min_fill_skips_partial_row_above_header(tmp_path):
    # A 3/5-filled metadata row sits above the real, fully-filled header.
    # By default the sparse row is mistaken for the header; header_min_fill
    # skips it (recording it) and anchors on the real header below.
    from openpyxl import Workbook

    p = tmp_path / "fill.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"
    ws.append(["Report", None, "v2", None, "2026"])  # 1: 60% filled junk
    ws.append(["Region", "Q1", "Q2", "Q3", "Q4"])  # 2: real header
    ws.append(["Alpha", 1, 2, 3, 4])  # 3: data
    ws.append(["Beta", 5, 6, 7, 8])  # 4: data
    wb.save(p)
    raw = reader.read_sheet(p, "All Data")

    # Default: the partial row wins (the pre-existing behaviour / the bug).
    default = extract.extract_table(raw)
    assert default.df.columns[0] == "Report"
    assert default.excel_rows == [2, 3, 4]

    # With a threshold the partial row is skipped as a pre-header note.
    ex = extract.extract_table(raw, header_min_fill=0.75)
    assert ex.df.columns == ["Region", "Q1", "Q2", "Q3", "Q4"]
    assert ex.excel_rows == [3, 4]
    assert [c["kind"] for c in ex.comments] == ["pre_header"]
    assert ex.comments[0]["excel_row"] == 1


def test_header_min_fill_allows_sparse_stacked_subheader(tmp_path):
    # Only the FIRST header row is fill-tested; a sparse stacked sub-header
    # below it must still be accepted as part of a multi-row header.
    ex = _extracted(
        tmp_path,
        fixtures.multi_header_book,
        "Wide",
        header_rows=2,
        header_min_fill=0.6,
    )
    assert ex.df.columns == ["Store", "Revenue Q1", "Revenue Q2"]


def test_header_min_fill_out_of_range(tmp_path):
    p = fixtures.simple_book(tmp_path / "s.xlsx")
    raw = reader.read_sheet(p, "Data")
    with pytest.raises(ValueError, match="header_min_fill"):
        extract.extract_table(raw, header_min_fill=1.5)


def _override_book(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "ov.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"
    ws.append(["Company XYZ", None, None, None, "sidebar"])  # 1
    ws.append(["Generated 2026", None, None, None, "note"])  # 2
    ws.append([None, None, None, None, None])  # 3 blank
    ws.append(["Region", "Q1", "Q2", "Q3", "misc"])  # 4 header
    ws.append(["Alpha", 1, 2, 3, "x"])  # 5
    ws.append(["Beta", 4, 5, 6, "y"])  # 6
    wb.save(p)
    return reader.read_sheet(p, "All Data")


def test_header_at_pins_the_header_row(tmp_path):
    raw = _override_book(tmp_path)
    ex = extract.extract_table(raw, header_at=4)
    assert ex.df.columns == ["Region", "Q1", "Q2", "Q3", "misc"]
    assert ex.excel_rows == [5, 6]
    assert any(e["event"] == "header_at" for e in ex.events)


def test_header_at_past_data_raises(tmp_path):
    raw = _override_book(tmp_path)
    with pytest.raises(EmptySheetError):
        extract.extract_table(raw, header_at=99)


def test_columns_override_forces_span(tmp_path):
    raw = _override_book(tmp_path)
    ex = extract.extract_table(raw, header_at=4, columns="A:D")
    assert ex.df.columns == ["Region", "Q1", "Q2", "Q3"]  # sidebar dropped
    ex = extract.extract_table(raw, header_at=4, columns="B:C")
    assert ex.df.columns == ["Q1", "Q2"]


def test_columns_override_clamped_to_extent(tmp_path):
    # An over-wide range is trimmed to the sheet's populated width rather than
    # manufacturing empty trailing columns.
    raw = _override_book(tmp_path)
    ex = extract.extract_table(raw, header_at=4, columns="A:ZZ")
    assert ex.df.width == 5


def test_columns_override_bad_spec(tmp_path):
    raw = _override_book(tmp_path)
    with pytest.raises(ValueError, match="column"):
        extract.extract_table(raw, columns="A:B:C")


def test_hidden_skipped_when_asked(tmp_path):
    ex = _extracted(
        tmp_path, fixtures.hidden_book, "Hidden", skip_hidden=True
    )
    assert ex.df.columns == ["A", "B"]
    assert ex.df.height == 2
    assert ex.excel_rows == [2, 4]


def test_empty_sheet_raises(tmp_path):
    p = fixtures.nasty_book(tmp_path / "n.xlsx")
    with pytest.raises(EmptySheetError):
        extract.extract_table(reader.read_sheet(p, "Empty"))


def test_mixed_type_column_becomes_text(tmp_path):
    ex = _extracted(tmp_path, fixtures.simple_book, "Data")
    assert ex.df["Revenue"].dtype == pl.Utf8


def test_units_column_is_integer(tmp_path):
    ex = _extracted(tmp_path, fixtures.nasty_book, "Report")
    assert ex.df["Units"].dtype == pl.Int64
    assert ex.df["Units"].to_list() == [10, 4, 7]


def test_integer_column_with_stray_float_stays_int(tmp_path):
    # An integer column where one cell is decimal-formatted (openpyxl -> float)
    # should downcast to Int64, matching pandas convert_dtypes.
    from openpyxl import Workbook

    p = tmp_path / "mix.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "M"
    ws.append(["n"])
    for v in (1, 2, 3.0, 4):
        ws.append([v])
    wb.save(p)
    ex = extract.extract_table(reader.read_sheet(p, "M"))
    assert ex.df["n"].dtype == pl.Int64
    assert ex.df["n"].to_list() == [1, 2, 3, 4]


def test_genuine_float_column_stays_float(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "f.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "F"
    ws.append(["n"])
    for v in (1.5, 2.0, 3.25):
        ws.append([v])
    wb.save(p)
    ex = extract.extract_table(reader.read_sheet(p, "F"))
    assert ex.df["n"].dtype == pl.Float64


def test_extraction_does_not_mutate_rawsheet(tmp_path):
    # Re-extracting the same RawSheet under different options must be safe
    # (kitchen_sink relies on this); the grid must not be mutated in place.
    p = fixtures.nasty_book(tmp_path / "n.xlsx")
    raw = reader.read_sheet(p, "Report")
    before = [list(r) for r in raw.grid]
    extract.extract_table(raw, sparse_rows="comment")
    assert raw.grid == before  # untouched
    ex2 = extract.extract_table(raw, sparse_rows="section")
    assert ex2.df.columns[0] == "section"
