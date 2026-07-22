"""Generated fixture workbooks. Each builder returns the saved path."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

ROWS = [
    ["Sydney", "01/02/2025", "$1,234"],
    ["Melbourne", "02/02/2025", "(500)"],
    ["Brisbane", "03/02/2025", "N/A"],
]


def simple_book(path: Path) -> Path:
    """One sheet 'Data', clean 1-row header, 3 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Store", "Date", "Revenue"])
    for r in ROWS:
        ws.append(r)
    wb.save(path)
    return path


def nasty_book(path: Path) -> Path:
    """Sheet 'Report': merged title banner (A1:D1), blank row, header,
    data with sparse section rows, plus a second empty sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "2025 Results — CONFIDENTIAL"
    ws.merge_cells("A1:D1")
    ws.append([])  # row 2 blank
    ws.append(["Store", "Date", "Units", "Revenue"])
    ws.append(["Northern Region"])  # sparse section row
    ws.append(["Sydney", "01/02/2025", 10, "$1,234"])
    ws.append(["Newcastle", "01/02/2025", 4, "$400"])
    ws.append(["Southern Region"])
    ws.append(["Melbourne", "02/02/2025", 7, "(500)"])
    wb.create_sheet("Empty")
    wb.save(path)
    return path


def multi_header_book(path: Path) -> Path:
    """Two header rows: 'Revenue' merged over Q1/Q2."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Wide"
    ws.append(["Store", "Revenue", None])
    ws.merge_cells("B1:C1")
    ws.append([None, "Q1", "Q2"])
    ws.append(["Sydney", 1, 2])
    ws.append(["Melbourne", 3, 4])
    wb.save(path)
    return path


def bloated_book(path: Path) -> Path:
    """Real table is 3x2 but a styled empty cell inflates the dimension
    to row 1,048,000. Must load fast with bounded memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bloat"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.append([3, 4])
    ws.cell(row=1_048_000, column=1).fill = PatternFill(
        "solid", fgColor="FFFF00"
    )
    wb.save(path)
    return path


def hidden_book(path: Path) -> Path:
    """Header + 3 data rows; row 3 (Excel) hidden, column C hidden."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden"
    ws.append(["A", "B", "Scratch"])
    ws.append([1, 2, "x"])
    ws.append([9, 9, "stale"])
    ws.append([3, 4, "y"])
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["C"].hidden = True
    wb.save(path)
    return path


def formula_book(path: Path) -> Path:
    """Formula cell with no cached value (never opened in Excel)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws.append(["A", "Total"])
    ws.append([1, "=SUM(A2:A2)"])
    wb.save(path)
    return path


def typed_book(path: Path) -> Path:
    """Columns exercising coercion: currency, percent, NA tokens,
    day-first dates, plain ints, free text."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"
    ws.append(["money", "pct", "when", "count", "note"])
    ws.append(["$1,234.50", "15%", "03/04/2025", 1, "hello"])
    ws.append(["(500)", "7.5%", "04/04/2025", 2, "world"])
    ws.append(["N/A", "-", "05/04/2025", 3, "-"])
    wb.save(path)
    return path
