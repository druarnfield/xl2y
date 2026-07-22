"""
excel_loader.py
===============
Load messy business Excel workbooks into clean pandas DataFrames, and
optionally persist each sheet's table to Parquet.

Assumptions / scope
-------------------
- At most ONE table per sheet.
- Title rows and comments (often merged cells) are detected, stripped from
  the data, and returned separately as annotations linked to the data row
  they precede.
- Only ``.xlsx`` / ``.xlsm`` / ``.xltx`` / ``.xltm`` are supported (openpyxl
  limitation). ``.xls`` / ``.xlsb`` raise a clear error asking for
  conversion.

Known ambiguity
---------------
A row merged across the full table width directly above the real header
(e.g. a merged "2024 Results" banner over ``Q1 | Q2 | Q3``) is
indistinguishable from a comment/title row and is classified as a comment.
If such a banner is actually the *top level of a multi-row header*, the
header will silently consume one data row. Inspect ``comments`` for
``kind == "merged"`` entries at ``excel_row`` just above your header if
you use ``header_rows > 1``.

Notes
-----
- Cell values are streamed in openpyxl's read_only mode; merged ranges and
  hidden-row/column flags come from a direct (streaming) parse of each
  sheet's XML. Nothing is fully materialised, so workbooks whose used range
  is inflated to row 1,048,576 by stray formatting ("dimension bloat") load
  in seconds with bounded memory.
- ``data_only=True`` returns Excel's *cached* formula results. Files written
  programmatically and never opened in Excel have no cache; such cells read
  as None. This loader detects that case and emits a warning.

Dependencies: openpyxl, pandas; pyarrow additionally for Parquet output.
"""

from __future__ import annotations

import json
import logging
import re
import xml.etree.ElementTree as ET
import zipfile
from bisect import bisect_right
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------- #
# Tunable heuristics
# --------------------------------------------------------------------------- #

#: A merged range overlapping at least this fraction of the table's visible
#: width is treated as a comment/title row -- but only if the row has no
#: other content outside the merge (so grouped multi-row headers are not
#: misclassified), and only if the merge's anchor actually holds a value
#: (formatting-only merges are ignored).
COMMENT_MERGE_WIDTH_FRACTION = 0.5

#: Minimum table width for the merged-comment heuristic to apply. In 1-2
#: column tables almost any merge spans "most of the width", so vertical
#: data merges would be eaten as comments without this guard.
COMMENT_MERGE_MIN_COLS = 3

#: A row with data in at most this fraction of the table's columns *and*
#: only non-numeric string content is treated as a comment/section row.
SPARSE_ROW_FRACTION = 0.25

#: Minimum table width for the sparse-row heuristic to apply.
SPARSE_ROW_MIN_COLS = 4

#: When coercing a text column to numeric/datetime, at least this fraction
#: of its non-null values must parse cleanly, otherwise the column is left
#: as text.
COERCE_MIN_FRACTION = 0.8

#: Case-insensitive tokens treated as missing values during coercion.
NA_TOKENS = frozenset(
    {
        "n/a",
        "na",
        "#n/a",
        "#value!",
        "#ref!",
        "#div/0!",
        "-",
        "--",
        "\u2013",
        "\u2014",
        "none",
        "null",
        "nil",
    }
)

_CURRENCY_RE = re.compile(
    r"^(?:a\$|au\$|aud|nz\$|us\$|usd|[$\u00a3\u20ac\u00a5])\s*", re.IGNORECASE
)

SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #


@dataclass
class SheetResult:
    """A cleaned table extracted from one worksheet."""

    sheet_name: str  #: Original sheet name as it appears in Excel.
    df: pd.DataFrame  #: Cleaned data with header applied.
    comments: list[dict] = field(default_factory=list)
    #: Each comment:
    #:   {"text": str, "excel_row": int, "kind": "merged"|"sparse",
    #:    "before_df_position": int|None}
    #: ``before_df_position`` is the *positional* index (``df.iloc``) of the
    #: first data row at or after the comment (None if the comment trails
    #: the table). It is a position, not an index label; they coincide only
    #: while the default RangeIndex is intact. Also stored in
    #: ``df.attrs["comments"]`` -- but pandas does not reliably preserve
    #: ``attrs`` through merge/concat; treat ``SheetResult`` as the source
    #: of truth.


class EmptySheetError(ValueError):
    """Raised when a sheet contains no data."""


class UnsupportedFormatError(ValueError):
    """Raised for spreadsheet formats openpyxl cannot open (.xls, .xlsb...)."""


def load_workbook_tables(
    path: str | Path,
    sheet_name: str | None = None,
    header_rows: int | dict[str, int] = 1,
    skip_empty_sheets: bool = True,
    sparse_rows: Literal["comment", "keep", "section"] = "comment",
    check_formula_cache: bool = True,
    coerce_values: bool = True,
    dayfirst: bool = True,
    skip_hidden: bool = False,
) -> dict[str, SheetResult]:
    """
    Load one sheet (if ``sheet_name`` given) or every worksheet.

    Chartsheets are skipped automatically (or raise a clear error if
    explicitly requested by name). Returns a dict keyed by the snake_case
    sheet name.

    Parameters
    ----------
    header_rows:
        Number of header rows, either a single int for all sheets or a dict
        mapping *original* sheet names to ints (missing sheets default to 1).
    sparse_rows:
        What to do with sparse text-only rows in wide tables (typically
        section labels like "Northern Region" in column A):

        - ``"comment"`` (default): strip them into ``comments`` with
          ``kind="sparse"`` (a WARNING is logged for each, since this can
          remove legitimate rows).
        - ``"keep"``: leave them in the data.
        - ``"section"``: strip them AND forward-fill their text into a new
          ``section`` column on the data rows beneath them -- usually what
          an analyst actually wants for sectioned reports.
    check_formula_cache:
        If True (default), scan the requested sheet(s) for formula cells
        whose cached value is missing and warn -- these read as None/NA.
    coerce_values:
        If True (default), text columns that are overwhelmingly numeric
        ("1,234", "$1,000", "(500)", "15%") or date-like are converted, and
        common NA tokens ("N/A", "-", "#REF!", ...) become missing values.
        Percentages become their fraction (``"15%"`` -> 0.15), matching how
        Excel stores real percentage cells. A column converts only when at
        least ``COERCE_MIN_FRACTION`` of its non-null values parse cleanly.
    dayfirst:
        Interpret ambiguous date strings like "03/04/2025" as day-first
        (3 April), per AU/UK convention. Set False for US-style month-first.
        Only affects string coercion; real Excel date cells are unambiguous.
    skip_hidden:
        If True, hidden rows and hidden columns are excluded from the table
        (hidden rows often hold stale scratch data). Default False.
    """
    path = Path(path)
    _check_format(path)
    meta = _sheet_meta(path)  # ordered: name -> (xml_target, is_chartsheet)
    worksheet_names = [n for n, (_, chart) in meta.items() if not chart]
    if sheet_name is not None:
        if sheet_name not in meta:
            raise KeyError(f"Sheet {sheet_name!r} not found. Available: {list(meta)}")
        if meta[sheet_name][1]:
            raise ValueError(
                f"Sheet {sheet_name!r} is a chartsheet and contains no cell data."
            )
        names = [sheet_name]
    else:
        skipped = [n for n, (_, chart) in meta.items() if chart]
        if skipped:
            logger.info("Skipping non-worksheet tab(s): %s", skipped)
        names = worksheet_names

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        results: dict[str, SheetResult] = {}
        with zipfile.ZipFile(path) as zf:
            for name in names:
                ws = wb[name]
                # Don't trust the declared dimension (stray formatting can
                # inflate it to XFD1048576); let openpyxl derive the extent
                # from the actual data while streaming.
                ws.reset_dimensions()
                (merged, hidden_rows, hidden_col_ivals, formulas, extent) = (
                    _sheet_structure(zf, meta[name][0])
                )
                n_header = (
                    header_rows.get(name, 1)
                    if isinstance(header_rows, dict)
                    else header_rows
                )
                try:
                    result = _extract_table(
                        name,
                        _read_grid(ws, extent),
                        merged_ranges=merged,
                        hidden_row_set=hidden_rows if skip_hidden else set(),
                        hidden_col_intervals=(hidden_col_ivals if skip_hidden else []),
                        header_rows=n_header,
                        sparse_rows=sparse_rows,
                        formula_cells=formulas if check_formula_cache else (),
                        coerce_values=coerce_values,
                        dayfirst=dayfirst,
                    )
                except EmptySheetError:
                    if skip_empty_sheets and sheet_name is None:
                        logger.info("Skipping empty sheet: %r", name)
                        continue
                    raise
                base = _snake_case(name)
                key, n = base, 2
                while key in results:  # e.g. "Q3 Sales" and "q3-sales"
                    key, n = f"{base}_{n}", n + 1
                results[key] = result
        return results
    finally:
        wb.close()


def load_sheet_table(
    path: str | Path,
    sheet_name: str,
    header_rows: int = 1,
    **kwargs: Any,
) -> SheetResult:
    """Extract the single table from one sheet.

    Thin wrapper around :func:`load_workbook_tables`; identical behavior,
    including raising :class:`EmptySheetError` for an empty sheet.
    """
    results = load_workbook_tables(
        path, sheet_name=sheet_name, header_rows=header_rows, **kwargs
    )
    return next(iter(results.values()))


def workbook_to_parquet(
    path: str | Path,
    out_dir: str | Path,
    overwrite: bool = False,
    **load_kwargs: Any,
) -> dict[str, dict]:
    """
    Load every table in the workbook and write one Parquet file per sheet.

    Returns a manifest dict keyed by snake_case sheet name::

        {
          "q3_sales": {
            "parquet_path": "/out/q3_sales.parquet",
            "sheet_name": "Q3 Sales",       # original Excel name
            "rows": 120, "columns": ["region", "revenue", ...],
            "comments": [...],              # same shape as SheetResult
          },
          ...
        }

    The comments and original sheet name are also embedded in each Parquet
    file's schema metadata under the key ``excel_loader`` (JSON), so the
    files remain self-describing::

        meta = pq.read_schema(p).metadata[b"excel_loader"]

    Columns that remain mixed-type ``object`` after loading are cast to
    string for Parquet compatibility (with a warning).

    ``load_kwargs`` are forwarded to :func:`load_workbook_tables`.
    """
    import pyarrow as pa
    import pyarrow.parquet as pq

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest: dict[str, dict] = {}
    for key, res in load_workbook_tables(path, **load_kwargs).items():
        target = out_dir / f"{key}.parquet"
        if target.exists() and not overwrite:
            raise FileExistsError(
                f"{target} already exists (pass overwrite=True to replace)."
            )

        df = res.df
        object_cols = [c for c in df.columns if df[c].dtype == object]
        if object_cols:
            logger.warning(
                "Sheet %r: mixed-type column(s) %s cast to string for Parquet.",
                res.sheet_name,
                object_cols,
            )
            df = df.copy()
            for c in object_cols:
                df[c] = (
                    df[c]
                    .map(lambda v: v if v is None or v is pd.NA else str(v))
                    .astype("string")
                )

        table = pa.Table.from_pandas(df, preserve_index=False)
        sidecar = {"sheet_name": res.sheet_name, "comments": res.comments}
        table = table.replace_schema_metadata(
            {
                **(table.schema.metadata or {}),
                b"excel_loader": json.dumps(sidecar).encode(),
            }
        )
        pq.write_table(table, target)

        manifest[key] = {
            "parquet_path": str(target),
            "sheet_name": res.sheet_name,
            "rows": len(df),
            "columns": list(map(str, df.columns)),
            "comments": res.comments,
        }
    return manifest


# --------------------------------------------------------------------------- #
# Internals
# --------------------------------------------------------------------------- #


def _check_format(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix in SUPPORTED_SUFFIXES:
        return
    if suffix in {".xls", ".xlsb"}:
        raise UnsupportedFormatError(
            f"{path.name}: {suffix} files cannot be read by openpyxl. "
            "Convert to .xlsx first (Excel: Save As; or `soffice --convert-to "
            "xlsx`; or pandas.read_excel with engine='xlrd'/'pyxlsb')."
        )
    raise UnsupportedFormatError(
        f"{path.name}: unrecognised extension {suffix!r}; expected one of "
        f"{sorted(SUPPORTED_SUFFIXES)}."
    )


def _snake_case(name: str) -> str:
    s = re.sub(r"[^\w]+", "_", name.strip())
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", s)  # camelCase boundaries
    s = re.sub(r"_+", "_", s).strip("_").lower() or "sheet"
    if not s[0].isalpha():  # "123_totals" is awkward in df.query/itertuples
        s = f"col_{s}"
    return s


def _is_empty(v: Any) -> bool:
    return v is None or str(v).strip() == ""


def _is_non_numeric_text(v: Any) -> bool:
    """True for strings that don't look like numbers ("1,234" is numeric)."""
    if not isinstance(v, str):
        return False
    try:
        float(v.strip().replace(",", ""))
        return False
    except ValueError:
        return True


def _localname(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _sheet_meta(path: Path) -> dict[str, tuple[str, bool]]:
    """Sheet order/type without loading the workbook.

    Returns an ordered mapping ``{sheet_name: (zip_member_path, is_chartsheet)}``
    parsed from ``xl/workbook.xml`` and its relationships.
    """
    RELS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(path) as zf:
        rels: dict[str, tuple[str, str]] = {}  # rId -> (target, type)
        root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        for rel in root:
            rels[rel.get("Id")] = (rel.get("Target"), rel.get("Type", ""))
        meta: dict[str, tuple[str, bool]] = {}
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        for el in wb_root.iter():
            if _localname(el.tag) != "sheet":
                continue
            rid = el.get(f"{{{RELS_NS}}}id")
            target, rel_type = rels.get(rid, ("", ""))
            # Targets may be absolute ("/xl/worksheets/sheet1.xml") or
            # relative to xl/ ("worksheets/sheet1.xml").
            member = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
            is_chart = "chartsheet" in rel_type or "chartsheets/" in member
            meta[el.get("name")] = (member, is_chart)
        return meta


def _sheet_structure(
    zf: zipfile.ZipFile, member: str
) -> tuple[
    list[tuple[int, int, int, int]],
    set[int],
    list[tuple[int, int]],
    list[tuple[int, int]],
    tuple[int, int],
]:
    """Merged ranges + hidden rows/columns via a streaming XML parse.

    Reading these through a full ``read_only=False`` workbook load would
    materialise every cell object -- catastrophic on dimension-bloated
    files -- when all we need is three kinds of small elements.

    Returns
    -------
    merged_ranges : list of (min_row, min_col, max_row, max_col), 1-indexed
    hidden_rows   : 0-indexed row numbers
    hidden_cols   : list of 0-indexed (min, max) inclusive intervals
                    (intervals, because "columns D:XFD hidden" is legal and
                    must not be expanded into 16k set entries)
    formula_cells : list of (row, col), 1-indexed
    data_extent   : (max_row, max_col), 1-indexed; (0, 0) if no cell holds
                    a value. Used to bound value iteration so styled-but-
                    empty ghost cells far below the table cost nothing.
    """
    merged: list[tuple[int, int, int, int]] = []
    hidden_rows: set[int] = set()
    hidden_col_intervals: list[tuple[int, int]] = []
    formulas: list[tuple[int, int]] = []
    max_row = max_col = 0
    truthy = {"1", "true"}
    # Cell/row `r` attributes are optional in OOXML; some writers omit them
    # and imply position, so track counters as we stream.
    cur_row = 0
    cur_col = 0
    cur_cell: tuple[int, int] | None = None
    with zf.open(member) as fh:
        for event, el in ET.iterparse(fh, events=("start", "end")):
            tag = _localname(el.tag)
            if event == "start":
                if tag == "row":
                    cur_row = int(el.get("r")) if el.get("r") else cur_row + 1
                    cur_col = 0
                elif tag == "c":
                    ref = el.get("r")
                    if ref:
                        cur_row, cur_col = coordinate_to_tuple(ref)
                    else:
                        cur_col += 1
                    cur_cell = (cur_row, cur_col)
                continue
            # end events
            if tag == "f":
                if cur_cell is not None:
                    formulas.append(cur_cell)
            elif tag in ("v", "is"):  # cell actually holds a value
                if cur_row > max_row:
                    max_row = cur_row
                if cur_col > max_col:
                    max_col = cur_col
            elif tag == "mergeCell":
                min_c, min_r, max_c, max_r = range_boundaries(el.get("ref"))
                merged.append((min_r, min_c, max_r, max_c))
            elif tag == "row":
                if el.get("hidden") in truthy:
                    hidden_rows.add(cur_row - 1)
                el.clear()  # rows dominate the file; free as we stream
            elif tag == "col":
                if el.get("hidden") in truthy:
                    hidden_col_intervals.append(
                        (int(el.get("min")) - 1, int(el.get("max")) - 1)
                    )
    return (merged, hidden_rows, hidden_col_intervals, formulas, (max_row, max_col))


def _read_grid(ws: Any, extent: tuple[int, int]) -> list[list[Any]]:
    """Stream cell values (read_only worksheet) into a trimmed grid.

    ``reset_dimensions()`` has been called, so iteration ends at the true
    data extent regardless of the declared dimension. Interior empty rows
    are yielded as placeholders (Excel row numbers are preserved); each row
    is trimmed of trailing empty cells, and runs of fully-empty rows are
    only flushed when a later non-empty row appears, so ghost regions cost
    one integer, not lists.
    """
    max_row, max_col = extent
    if max_row == 0:
        return []
    grid: list[list[Any]] = []
    pending_empty = 0
    for row in ws.iter_rows(max_row=max_row, max_col=max_col, values_only=True):
        vals = list(row)
        while vals and _is_empty(vals[-1]):
            vals.pop()
        if not vals:
            pending_empty += 1
        else:
            if pending_empty:
                grid.extend([] for _ in range(pending_empty))
                pending_empty = 0
            grid.append(vals)
    return grid  # trailing empty rows deliberately dropped


def _extract_table(
    title: str,
    grid: list[list[Any]],
    merged_ranges: list[tuple[int, int, int, int]],
    hidden_row_set: set[int],
    hidden_col_intervals: list[tuple[int, int]],
    header_rows: int = 1,
    sparse_rows: str = "comment",
    formula_cells: tuple | list = (),
    coerce_values: bool = True,
    dayfirst: bool = True,
) -> SheetResult:
    # ------------------------------------------------------------------ #
    # 1. Merged cells only store a value in the top-left anchor; propagate
    #    it across the *full* merged rectangle (both horizontally and
    #    vertically) so category labels merged down a column, and comments
    #    merged across several rows, are visible. Rows are padded on
    #    demand so merges reaching past a trimmed row still propagate.
    # ------------------------------------------------------------------ #
    if not grid:
        raise EmptySheetError(f"Sheet {title!r} is empty")

    # Snapshot which formula cells are uncached BEFORE propagation, so a
    # merged-range copy can't mask a missing cache value. (Counted against
    # the table's bounding box later.)
    uncached_coords = {
        (fr - 1, fc - 1)
        for fr, fc in formula_cells
        if fr - 1 >= len(grid)
        or fc - 1 >= len(grid[fr - 1])
        or grid[fr - 1][fc - 1] is None
    }

    propagated: set[tuple[int, int]] = set()  # 0-indexed copies (not anchors)
    for min_r, min_c, max_r, max_c in merged_ranges:
        if min_r - 1 >= len(grid) or min_c - 1 >= len(grid[min_r - 1]):
            continue
        anchor = grid[min_r - 1][min_c - 1]
        if anchor is None:
            continue
        for r in range(min_r - 1, min(max_r, len(grid))):
            if len(grid[r]) < max_c:  # pad trimmed rows inside the merge
                grid[r].extend([None] * (max_c - len(grid[r])))
            for c in range(min_c - 1, max_c):
                if (r, c) != (min_r - 1, min_c - 1):
                    if grid[r][c] is None:
                        grid[r][c] = anchor
                    propagated.add((r, c))

    # ------------------------------------------------------------------ #
    # 2. Bounding box: trim fully-empty leading/trailing rows and columns,
    #    and (optionally) drop hidden rows/columns.
    # ------------------------------------------------------------------ #
    hidden_rows = hidden_row_set
    grid_cols = max((len(r) for r in grid), default=0)
    hidden_cols = {
        c
        for lo, hi in hidden_col_intervals
        for c in range(max(lo, 0), min(hi, grid_cols - 1) + 1)
    }
    if hidden_rows or hidden_cols:
        logger.info(
            "Sheet %r: skipping %d hidden row(s), %d hidden column(s).",
            title,
            len(hidden_rows),
            len(hidden_cols),
        )

    def visible_rows() -> list[int]:
        return [r for r in range(len(grid)) if r not in hidden_rows]

    row_has = {r: any(not _is_empty(v) for v in grid[r]) for r in visible_rows()}
    data_rows = [r for r, has in row_has.items() if has]
    if not data_rows:
        raise EmptySheetError(f"Sheet {title!r} has no data")
    top, bottom = data_rows[0], data_rows[-1]

    # Column extent: notes typed/merged off to the side of the table would
    # inflate a naive bounding box (and break the width-based comment
    # heuristics). Split the columns into contiguous runs separated by
    # fully-empty columns and take the run containing the most data as THE
    # table. Caveat: a table with an intentionally all-empty spacer column
    # inside it will be truncated at the spacer.
    # NOTE: hidden columns still count here -- they are excluded from the
    # output via ``table_cols`` below, but must not act as run separators
    # (a hidden column in the middle of a table would otherwise split it).
    n_cols = max(len(grid[r]) for r in data_rows)
    col_counts = [
        sum(
            1
            for r in range(top, bottom + 1)
            if r not in hidden_rows and len(grid[r]) > c and not _is_empty(grid[r][c])
        )
        for c in range(n_cols)
    ]
    runs: list[tuple[int, int]] = []  # (start, end) inclusive, non-empty cols
    start = None
    for c, count in enumerate(col_counts + [0]):
        if count and start is None:
            start = c
        elif not count and start is not None:
            runs.append((start, c - 1))
            start = None
    left, right = max(runs, key=lambda run: sum(col_counts[run[0] : run[1] + 1]))
    table_cols = [c for c in range(left, right + 1) if c not in hidden_cols]
    width = len(table_cols)
    if len(runs) > 1:
        logger.info(
            "Sheet %r: multiple column blocks found; using columns %d-%d as "
            "the table and ignoring the rest.",
            title,
            left + 1,
            right + 1,
        )

    # ------------------------------------------------------------------ #
    # 3. Merged-comment detection. A merge counts as a comment/title block
    #    when its anchor holds a value (formatting-only merges are common
    #    and meaningless), it overlaps most of the table's visible width,
    #    AND the rows it covers have no content outside the merge (this
    #    preserves grouped headers like "Revenue" merged over 4 of 6
    #    columns next to other header cells). Multi-row comment merges
    #    flag every covered row. Skipped entirely for very narrow tables,
    #    where vertical data merges would otherwise qualify.
    # ------------------------------------------------------------------ #
    merged_comment_rows: dict[int, tuple[int, int]] = {}  # row -> anchor (r, c)
    if width >= COMMENT_MERGE_MIN_COLS:
        for min_r, min_c, max_r, max_c in merged_ranges:
            if min_r - 1 >= len(grid) or min_c - 1 >= len(grid[min_r - 1]):
                continue
            if _is_empty(grid[min_r - 1][min_c - 1]):  # formatting-only merge
                continue
            overlap = sum(1 for c in table_cols if min_c - 1 <= c <= max_c - 1)
            if overlap < width * COMMENT_MERGE_WIDTH_FRACTION:
                continue
            span_rows = [
                r
                for r in range(min_r - 1, min(max_r, len(grid)))
                if r not in hidden_rows
            ]
            clean = all(
                (_is_empty(grid[r][c]) if c < len(grid[r]) else True)
                for r in span_rows
                for c in table_cols
                if not (min_c - 1 <= c <= max_c - 1)
            )
            if clean:
                for r in span_rows:
                    merged_comment_rows.setdefault(r, (min_r - 1, min_c - 1))

    # ------------------------------------------------------------------ #
    # 4. Classify rows: comments first, THEN headers, then body. This way
    #    a merged title above the real header no longer swallows it.
    #    All-empty body rows are dropped here (not via df.dropna later),
    #    so df positions computed below are final and comment links stay
    #    correct.
    # ------------------------------------------------------------------ #
    header: list[list[Any]] = []
    body_rows: list[tuple[int, list[Any], str | None]] = []
    #                     (excel_row_1idx, values, section)
    raw_comments: list[tuple[int, str, str]] = []  # (excel_row_1idx, text, kind)
    seen_merge_anchors: set[tuple[int, int]] = set()  # (row, col) per block
    current_section: str | None = None

    def row_text(r: int, values: list[Any]) -> str:
        # Propagated merge copies are excluded, so no lossy de-dupe is
        # needed and distinct cells with identical text are preserved.
        parts = []
        for c, v in zip(table_cols, values):
            if _is_empty(v):
                continue
            if (r, c) in propagated:
                continue
            parts.append(str(v).strip())
        return " ".join(parts)

    for r in range(top, bottom + 1):
        if r in hidden_rows:
            continue
        values = [grid[r][c] if c < len(grid[r]) else None for c in table_cols]

        if r in merged_comment_rows:
            anchor = merged_comment_rows[r]
            if anchor not in seen_merge_anchors:  # one entry per merge block
                seen_merge_anchors.add(anchor)
                # Propagated copies are skipped, so the merged value appears
                # once; other (unmerged) cells on the row are still included.
                text = row_text(r, values)
                if text:  # never record blank comments
                    raw_comments.append((anchor[0] + 1, text, "merged"))
            continue

        # Cells whose only content was carried in by a vertical merge are
        # not "real" content for classification purposes: a row consisting
        # solely of a propagated category label is an empty row, not data
        # and not a sparse comment.
        own_non_null = [
            v
            for c, v in zip(table_cols, values)
            if not _is_empty(v) and (r, c) not in propagated
        ]
        non_null = [v for v in values if not _is_empty(v)]

        if not non_null:
            # Fully-empty rows are never headers, comments, or data. (This
            # matters when a blank row sits between a title and the header.)
            continue

        # Sparse text-only row in a wide table -> section label / comment.
        if (
            sparse_rows in ("comment", "section")
            and len(header) >= header_rows  # never applies to header rows
            and width >= SPARSE_ROW_MIN_COLS
            and 0 < len(own_non_null) <= max(1, int(width * SPARSE_ROW_FRACTION))
            and len(own_non_null) == len(non_null)  # no propagated data cells
            and all(_is_non_numeric_text(v) for v in own_non_null)
        ):
            text = row_text(r, values)
            if sparse_rows == "section":
                current_section = text
                logger.info(
                    "Sheet %r: row %d used as section label %r.",
                    title,
                    r + 1,
                    text,
                )
            else:
                logger.warning(
                    "Sheet %r: treating sparse row %d as a comment "
                    "(text: %.80r). Pass sparse_rows='keep' to keep such "
                    "rows as data, or sparse_rows='section' to forward-fill "
                    "them into a section column.",
                    title,
                    r + 1,
                    text,
                )
                raw_comments.append((r + 1, text, "sparse"))
            continue

        if len(header) < header_rows:
            header.append(values)
            continue

        if own_non_null:  # drop rows with no content of their own
            body_rows.append((r + 1, values, current_section))

    if len(header) < header_rows:
        raise ValueError(
            f"Sheet {title!r}: requested header_rows={header_rows} but only "
            f"{len(header)} non-comment row(s) available."
        )

    # Formula-cache warning, scoped to the table's bounding box and based
    # on pre-propagation values.
    uncached = sum(
        1
        for (r, c) in uncached_coords
        if top <= r <= bottom and c in table_cols and r not in hidden_rows
    )
    if uncached:
        logger.warning(
            "Sheet %r: %d formula cell(s) have no cached value and will read "
            "as missing. Open and save the file in Excel to populate them.",
            title,
            uncached,
        )

    # ------------------------------------------------------------------ #
    # 5. Build the DataFrame and link comments to the first data row at or
    #    after each comment (binary search; body_rows is already sorted).
    # ------------------------------------------------------------------ #
    columns = _build_columns(header, width)
    df = pd.DataFrame([v for _, v, _ in body_rows], columns=columns)
    if sparse_rows == "section" and any(s is not None for _, _, s in body_rows):
        section_col = _unique_name("section", set(columns))
        df.insert(0, section_col, [s for _, _, s in body_rows])
    df = df.convert_dtypes()
    if coerce_values and not df.empty:
        df = _coerce_columns(df, title, dayfirst=dayfirst)

    body_excel_rows = [er for er, _, _ in body_rows]
    comments = []
    for excel_row, text, kind in sorted(raw_comments):
        pos = bisect_right(body_excel_rows, excel_row - 1)
        comments.append(
            {
                "text": text,
                "excel_row": excel_row,
                "kind": kind,
                "before_df_position": pos if pos < len(body_excel_rows) else None,
            }
        )

    result = SheetResult(sheet_name=title, df=df, comments=comments)
    df.attrs["comments"] = comments  # convenience copy; see SheetResult docs
    return result


def _unique_name(base: str, taken: set[str]) -> str:
    name, n = base, 2
    while name in taken:
        name, n = f"{base}_{n}", n + 1
    return name


def _build_columns(header: list[list[Any]], width: int) -> list[str]:
    if not header:
        return [f"col_{i}" for i in range(width)]
    parts_per_col: list[list[str]] = []
    for c in range(width):
        parts, last = [], None
        for row in header:
            v = row[c] if c < len(row) else None
            v = str(v).strip() if v is not None and str(v).strip() else None
            if v is not None and v != last:
                parts.append(v)
                last = v
        parts_per_col.append(parts)
    names: list[str] = []
    taken: set[str] = set()
    for i, parts in enumerate(parts_per_col):
        base = _snake_case(" ".join(parts)) if parts else f"col_{i}"
        # Check generated suffixes against ALL names so "value, value,
        # value_2" cannot collide into duplicate columns.
        name = _unique_name(base, taken)
        taken.add(name)
        names.append(name)
    return names


# --------------------------------------------------------------------------- #
# Value coercion
# --------------------------------------------------------------------------- #


def _parse_numberish(v: Any) -> float | None:
    """Parse '1,234', '$1,000.50', 'A$ 2,000', '(500)', '15%'; None if not."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):  # accounting negatives
        negative, s = True, s[1:-1].strip()
    if s.startswith("-"):
        negative, s = True, s[1:].strip()
    s = _CURRENCY_RE.sub("", s)
    percent = s.endswith("%")
    if percent:
        s = s[:-1].strip()
    s = s.replace(",", "").replace("\u00a0", "").replace(" ", "")
    if not s:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    if percent:
        n /= 100.0
    return -n if negative else n


def _coerce_columns(df: pd.DataFrame, sheet: str, dayfirst: bool) -> pd.DataFrame:
    """Column-level coercion of text columns to numeric/datetime + NA tokens.

    A column converts only when >= COERCE_MIN_FRACTION of its non-null
    values parse; the stragglers become NA (with a warning). Otherwise the
    column is left untouched.
    """
    df = df.copy()
    for col in df.columns:
        if not (df[col].dtype == object or str(df[col].dtype) == "string"):
            continue
        s = df[col]
        # NA tokens first (applies even if the column stays text).
        is_na_token = s.map(
            lambda v: isinstance(v, str) and v.strip().lower() in NA_TOKENS
        )
        if is_na_token.any():
            s = s.mask(is_na_token, pd.NA)
        values = s.dropna()
        if values.empty:
            df[col] = s
            continue

        # --- numeric ---------------------------------------------------- #
        parsed = values.map(_parse_numberish)
        ok = parsed.notna()
        if ok.mean() >= COERCE_MIN_FRACTION:
            failed = int((~ok).sum())
            if failed:
                logger.warning(
                    "Sheet %r: column %r coerced to numeric; %d value(s) "
                    "did not parse and became NA (e.g. %r).",
                    sheet,
                    col,
                    failed,
                    values[~ok].iloc[0],
                )
            out = pd.Series(parsed, index=values.index, dtype="Float64")
            df[col] = out.reindex(df.index)
            if (df[col].dropna() % 1 == 0).all():
                df[col] = df[col].astype("Int64")
            continue

        # --- datetime ---------------------------------------------------- #
        strings = values.map(lambda v: v if isinstance(v, str) else None).dropna()
        if not strings.empty and len(strings) == len(values):
            dt = pd.to_datetime(
                strings, errors="coerce", dayfirst=dayfirst, format="mixed"
            )
            ok = dt.notna()
            # Guard: to_datetime happily parses bare numbers/years; require
            # at least one separator or month word to call it a date column.
            looks_datey = strings.str.contains(
                r"[-/:.]|\d\s+\w|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec",
                case=False,
                regex=True,
            )
            if ok.mean() >= COERCE_MIN_FRACTION and looks_datey.mean() >= 0.5:
                failed = int((~ok).sum())
                if failed:
                    logger.warning(
                        "Sheet %r: column %r coerced to datetime; %d value(s) "
                        "did not parse and became NA.",
                        sheet,
                        col,
                        failed,
                    )
                logger.info(
                    "Sheet %r: column %r parsed as dates (dayfirst=%s).",
                    sheet,
                    col,
                    dayfirst,
                )
                df[col] = dt.reindex(df.index)
                continue

        df[col] = s  # keep as text (NA tokens still applied)
    return df


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(description="Excel -> clean DataFrame(s)")
    parser.add_argument("path")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: all)")
    parser.add_argument("--header-rows", type=int, default=1)
    parser.add_argument(
        "--sparse-rows",
        choices=["comment", "keep", "section"],
        default="comment",
        help="Sparse text-only rows: strip as comments "
        "(default), keep as data, or forward-fill into "
        "a 'section' column",
    )
    parser.add_argument(
        "--no-formula-check", action="store_true", help="Skip the uncached-formula scan"
    )
    parser.add_argument(
        "--no-coerce",
        action="store_true",
        help="Do not coerce '$1,234' / '15%%' / NA tokens / date strings",
    )
    parser.add_argument(
        "--monthfirst",
        action="store_true",
        help="Parse ambiguous dates as month-first (US) "
        "instead of day-first (AU default)",
    )
    parser.add_argument(
        "--skip-hidden", action="store_true", help="Exclude hidden rows and columns"
    )
    parser.add_argument(
        "--include-empty-sheets",
        action="store_true",
        help="Raise on empty sheets instead of skipping",
    )
    parser.add_argument(
        "--parquet-dir",
        default=None,
        metavar="DIR",
        help="Also write one Parquet file per sheet to DIR and print the manifest",
    )
    parser.add_argument(
        "--overwrite", action="store_true", help="Overwrite existing Parquet files"
    )
    args = parser.parse_args()

    kwargs: dict[str, Any] = dict(
        sheet_name=args.sheet,
        header_rows=args.header_rows,
        sparse_rows=args.sparse_rows,
        check_formula_cache=not args.no_formula_check,
        coerce_values=not args.no_coerce,
        dayfirst=not args.monthfirst,
        skip_hidden=args.skip_hidden,
        skip_empty_sheets=not args.include_empty_sheets,
    )

    if args.parquet_dir:
        manifest = workbook_to_parquet(
            args.path, args.parquet_dir, overwrite=args.overwrite, **kwargs
        )
        print(json.dumps(manifest, indent=2, default=str))
    else:
        for key, res in load_workbook_tables(args.path, **kwargs).items():
            print(
                f"\n=== {key} (sheet: {res.sheet_name!r}) "
                f"{res.df.shape[0]} rows x {res.df.shape[1]} cols ==="
            )
            print(res.df.head(10).to_string())
            for c in res.comments:
                print(
                    f"  [{c['kind']} comment @ Excel row {c['excel_row']}, "
                    f"before df position {c['before_df_position']}]: "
                    f"{c['text']}"
                )
